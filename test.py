from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from time import sleep
from openpyxl import load_workbook


dataSheet = 'challenge.xlsx'
openingSpreadsheet = load_workbook( dataSheet)
selectedWorksheet = openingSpreadsheet ['Sheet1']

opemBrowser = webdriver.Edge()
opemBrowser.get('https://rpachallenge.com/')

startChallenge = opemBrowser.find_element(By.XPATH, '/html/body/app-root/div[2]/app-rpa1/div/div[1]/div[6]/button' )
startChallenge.click()
sleep(2)

for line in range(2, len(selectedWorksheet['A']) + 1 ):
    First_Name = selectedWorksheet[f'A{line}'].value
    Last_Name = selectedWorksheet[f'B{line}'].value
    Company_Name = selectedWorksheet[f'C{line}'].value
    Role_in_Company = selectedWorksheet[f'D{line}'].value
    Address = selectedWorksheet[f'E{line}'].value
    Email = selectedWorksheet[f'F{line}'].value
    Phone_Number = selectedWorksheet[f'G{line}'].value
    if First_Name != None:
        populate = opemBrowser.find_element(By.XPATH, "//input[@ng-reflect-name='labelLastName']")
        populate.clear()
        populate.send_keys(Last_Name)
        sleep(0.1)
        populate = opemBrowser.find_element(By.XPATH, "//input[@ng-reflect-name='labelFirstName']")
        populate.clear()
        populate.send_keys(First_Name)
        sleep(0.1)
        populate = opemBrowser.find_element(By.XPATH, "//input[@ng-reflect-name='labelCompanyName']")
        populate.clear()
        populate.send_keys(Company_Name)
        sleep(0.1)
        populate = opemBrowser.find_element(By.XPATH, "//input[@ng-reflect-name='labelRole']")
        populate.clear()
        populate.send_keys(Role_in_Company)
        sleep(0.1)
        populate = opemBrowser.find_element(By.XPATH, "//input[@ng-reflect-name='labelAddress']")
        populate.clear()
        populate.send_keys(Address)
        sleep(0.1)
        populate = opemBrowser.find_element(By.XPATH, "//input[@ng-reflect-name='labelEmail']")
        populate.clear()
        populate.send_keys(Email)
        sleep(0.1)
        populate = opemBrowser.find_element(By.XPATH, "//input[@ng-reflect-name='labelPhone']")
        populate.clear()
        populate.send_keys(Phone_Number)

        button = opemBrowser.find_element(By.XPATH, '/html/body/app-root/div[2]/app-rpa1/div/div[2]/form/input')
        button.click()       
sleep(3)





# from selenium import webdriver
# from selenium.webdriver.common.keys import Keys
# from selenium.webdriver.common.by import By
# from time import sleep
# from openpyxl import load_workbook

# dataSheet = 'challenge.xlsx'
# openingSpreadsheet = load_workbook(dataSheet)
# selectedWorksheet = openingSpreadsheet['Sheet1']

# opemBrowser = webdriver.Edge()
# opemBrowser.get('https://rpachallenge.com/')

# sleep(2)

# # Inicializa um array vazio para armazenar os dados
# data_array = []

# for line in range(2, len(selectedWorksheet['A']) + 1):
#     First_Name = selectedWorksheet[f'A{line}'].value
#     Last_Name = selectedWorksheet[f'B{line}'].value
#     Company_Name = selectedWorksheet[f'C{line}'].value
#     Role_in_Company = selectedWorksheet[f'D{line}'].value
#     Address = selectedWorksheet[f'E{line}'].value
#     Email = selectedWorksheet[f'F{line}'].value
#     Phone_Number = selectedWorksheet[f'G{line}'].value
    
#     # Verifica se o First_Name não é None antes de adicionar ao array
#     if First_Name is not None:
#         # Adiciona os valores em formato de dicionário ao array
#         data_array.append({
#             'First_Name': First_Name,
#             'Last_Name': Last_Name,
#             'Company_Name': Company_Name,
#             'Role_in_Company': Role_in_Company,
#             'Address': Address,
#             'Email': Email,
#             'Phone_Number': Phone_Number
#         })

# # Imprime o array resultante
# print(data_array)

# input('Enter')
